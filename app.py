import streamlit as st
import pandas as pd
from io import BytesIO
import openpyxl
from openpyxl.styles import Alignment, PatternFill, Border, Side
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email.utils import formataddr
from email import encoders

# --- EMAIL CONFIGURATION ---
SENDER_EMAIL = "atharvaujoshi@gmail.com"
SENDER_NAME = "Spydarr Summary to Report" 
APP_PASSWORD = "nybl zsnx zvdw edqr"  

def send_email(recipient_email, excel_data, filename):
    try:
        recipient_name = recipient_email.split('@')[0].replace('.', ' ').title()
        msg = MIMEMultipart()
        msg['From'] = formataddr((SENDER_NAME, SENDER_EMAIL))
        msg['To'] = recipient_email
        msg['Subject'] = "Spydarr Summary to Report"
        body = f"Dear {recipient_name},\n\nPlease find the attached professional Market Report.\n\nRegards,\nAtharva Joshi"
        msg.attach(MIMEText(body, 'plain'))
        part = MIMEBase('application', 'octet-stream')
        part.set_payload(excel_data)
        encoders.encode_base64(part)
        part.add_header('Content-Disposition', f"attachment; filename={filename}")
        msg.attach(part)
        server = smtplib.SMTP('smtp.gmail.com', 587)
        server.starttls()
        server.login(SENDER_EMAIL, APP_PASSWORD)
        server.send_message(msg)
        server.quit()
        return True
    except Exception as e:
        st.error(f"Error sending email: {e}")
        return False

# --- STREAMLIT UI SETUP ---
st.set_page_config(page_title="Property Report Tool", layout="wide")
st.title("🏙️Spydarr's Summary to Report")

# Fixed Indentation and String Closing here
st.markdown("""
        <span style='background-color: #FFFF00; padding: 2px 8px; border-radius: 4px; border: 1px solid #E6E600; font-size: 0.9em; color: black;'>
            <u><strong>NOTE :-</strong> Before Uploading the Summary Cross-Check it.</u>
    <strong>Need the base summary first?</strong> 
    <a href="https://spydarr.streamlit.app/" target="_blank">To get the summary click here: Spydarr Dashboard</a>
</div>
""", unsafe_allow_html=True)

uploaded_file = st.file_uploader("Upload your Excel file", type=['xlsx'])

if uploaded_file:
    try:
        xls = pd.ExcelFile(uploaded_file)
        target_sheet = next((s for s in xls.sheet_names if s.lower() == 'summary'), None)
        
        if not target_sheet:
            st.error("Could not find a sheet named 'summary' or 'Summary'.")
        else:
            df = pd.read_excel(xls, sheet_name=target_sheet)

            # --- DATA PROCESSING ---
            df['Location'] = df['Location'].ffill()
            df['Property'] = df['Property'].ffill()
            df['Total Count'] = df['Total Count'].ffill()
            df['Last Completion Date'] = pd.to_datetime(df['Last Completion Date'], dayfirst=True)
            df['weighted_apr_sum'] = df['Average of APR'] * df['Count of Property']

            report_df = df.groupby(['Location', 'Property', 'Last Completion Date', 'Configuration']).agg({
                'Carpet Area(SQ.FT)': ['min', 'max'],
                'Min. APR': 'min',
                'Max APR': 'max',
                'weighted_apr_sum': 'sum',
                'Count of Property': 'sum',
                'Total Count': 'first'
            }).reset_index()

            report_df.columns = ['Location', 'Property', 'Last Completion Date', 'Configuration', 
                                 'min_carpet', 'max_carpet', 'Min APR', 'Max APR', 
                                 'temp_sum_apr', 'Count of Property', 'Total Count']

            report_df['Carpet Area(SQ.FT)'] = report_df.apply(
                lambda x: str(int(round(x['min_carpet']))) if round(x['min_carpet']) == round(x['max_carpet']) 
                else f"{int(round(x['min_carpet']))}-{int(round(x['max_carpet']))}", axis=1
            )

            report_df['Average of APR'] = (report_df['temp_sum_apr'] / report_df['Count of Property']).fillna(0)
            
            numeric_cols = ['Min APR', 'Max APR', 'Average of APR', 'Count of Property', 'Total Count']
            for col in numeric_cols:
                report_df[col] = report_df[col].fillna(0).round(0).astype(int)
            
            report_df['Last Completion Date'] = report_df['Last Completion Date'].dt.strftime('%b-%y')
            final_df = report_df[['Location', 'Property', 'Last Completion Date', 'Configuration', 
                                  'Carpet Area(SQ.FT)', 'Min APR', 'Max APR', 'Average of APR', 
                                  'Count of Property', 'Total Count']]

            # --- EXCEL STYLING ---
            output = BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                final_df.to_excel(writer, index=False, sheet_name='Report')
                ws = writer.book['Report']
                
                center_align = Alignment(horizontal='center', vertical='center')
                thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                                     top=Side(style='thin'), bottom=Side(style='thin'))
                colors = ["A2D2FF", "FFD6A5", "CAFFBF", "FDFFB6", "FFADAD", "BDB2FF", "9BF6FF"]
                
                last_row = len(final_df) + 1
                last_col = len(final_df.columns)

                # 1. Apply Global Alignment and Borders
                for r in range(1, last_row + 1): 
                    for c in range(1, last_col + 1):
                        cell = ws.cell(row=r, column=c)
                        cell.alignment = center_align 
                        cell.border = thin_border

                # 2. Logic for Merging LOCATION (Column 1)
                current_loc, start_row_loc = None, 2
                for row_num in range(2, last_row + 2):
                    row_loc = ws.cell(row=row_num, column=1).value 
                    if row_loc != current_loc or row_num == last_row + 1:
                        if current_loc is not None:
                            end_row_loc = row_num - 1
                            if end_row_loc > start_row_loc:
                                ws.merge_cells(start_row=start_row_loc, start_column=1, end_row=end_row_loc, end_column=1)
                        start_row_loc, current_loc = row_num, row_loc

                # 3. Logic for Merging PROPERTY (Col 2), TOTAL COUNT (Col 10), and COLORING
                current_prop, start_row_prop, color_idx = None, 2, 0
                for row_num in range(2, last_row + 2):
                    row_prop = ws.cell(row=row_num, column=2).value 
                    if row_prop != current_prop or row_num == last_row + 1:
                        if current_prop is not None:
                            end_row_prop = row_num - 1
                            fill = PatternFill(start_color=colors[color_idx % len(colors)], fill_type="solid")
                            
                            for r_fill in range(start_row_prop, end_row_prop + 1):
                                for c_fill in range(1, last_col + 1): 
                                    ws.cell(row=r_fill, column=c_fill).fill = fill
                            
                            if end_row_prop > start_row_prop:
                                ws.merge_cells(start_row=start_row_prop, start_column=2, end_row=end_row_prop, end_column=2)
                                ws.merge_cells(start_row=start_row_prop, start_column=last_col, end_row=end_row_prop, end_column=last_col)
                            
                            color_idx += 1
                        start_row_prop, current_prop = row_num, row_prop
                
                for col in ws.columns: 
                    ws.column_dimensions[col[0].column_letter].width = 20

            file_content = output.getvalue()

            st.sidebar.divider()
            st.sidebar.header("📧 Email Report")
            recipient = st.sidebar.text_input("Recipient Name", placeholder="firstname.lastname")
            
            if st.sidebar.button("Send to Email") and recipient:
                full_email = f"{recipient.strip().lower()}@beyondwalls.com"
                with st.spinner(f'Sending to {full_email}...'):
                    if send_email(full_email, file_content, "Spydarr_Summary_to_Report.xlsx"):
                        st.sidebar.success(f"Report sent to {full_email}")
            

    except Exception as e:
        st.error(f"Error: {e}")
